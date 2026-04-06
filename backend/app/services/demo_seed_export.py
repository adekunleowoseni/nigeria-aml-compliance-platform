"""
Build a multi-sheet .xlsx of static demo / seed reference data (no DB writes).

Sheets: README, OTC branch reference, standard AML seed, showcase 12 tracks, ingest flagship,
temporal profiles, temporal scenario codes.
"""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from uuid import uuid4

from app.models.transaction import TransactionResponse
from app.services.demo_showcase import seed_high_risk_showcase
from app.services.demo_standard_seed import run_standard_demo_transaction_sequence
from app.services.otc_branch_reference_seed import export_reference_table_json_ready
from app.services.temporal_simulation import DEFAULT_PROFILES

# Stable anchor for exported datetimes (matches platform docs / sample era).
_EXPORT_REFERENCE_NOW = datetime(2026, 4, 3, 12, 0, 0)

_TEMPORAL_SCENARIO_CODES: tuple[str, ...] = (
    "SMURFING_FAN_IN",
    "SMURFING_INTEGRATION_OUT",
    "LAYERING_PASS_THROUGH",
    "CASH_PROFILE_MISMATCH",
    "CASH_PROFILE_FOLLOWUP_OUT",
    "STRUCTURING",
    "STRUCTURING_INTEGRATION_OUT",
    "VELOCITY_BURST",
    "VELOCITY_SETTLEMENT_OUT",
    "WIRE_SPIKE",
    "WIRE_SPIKE_OUT",
    "ROUND_TRIP",
    "ROUND_TRIP_FOLLOWOUT",
    "CHANNEL_ANOMALY",
    "CHANNEL_ANOMALY_OUT",
)


def _flagship_template(now: datetime) -> TransactionResponse:
    return TransactionResponse(
        id=str(uuid4()),
        customer_id="DEMO-PERSON-ADESANYA",
        amount=55_000_000.00,
        currency="NGN",
        transaction_type="wire",
        narrative=(
            "FCMB SWIFT — Federal Ministry of Works refund referenced in memo; "
            "beneficiary individual salary account (PEP-style review required)"
        ),
        counterparty_id="NG-FMW-REFUND",
        counterparty_name="Federal Ministry of Works & Housing",
        status="received",
        created_at=now,
        metadata={
            "profile": "civil_servant_ippis",
            "sender_bank": "First City Monument Bank",
            "pep_flag": True,
            "counterparty_type": "government_entity",
            "simulation_scenario": "FLAGSHIP_DEMO",
        },
    )


def _transactions_to_sheet(ws, txns: List[TransactionResponse], sheet_title: str = "") -> None:
    ws.cell(1, 1, f"Sheet: {sheet_title}")
    ws.cell(1, 1).font = Font(bold=True)
    if not txns:
        ws.cell(2, 1, "(no rows)")
        return
    rows_out: List[Dict[str, Any]] = []
    for t in txns:
        d = t.model_dump(mode="json")
        md = d.pop("metadata", None)
        d["metadata_json"] = json.dumps(md, ensure_ascii=False) if md else ""
        rows_out.append(d)
    keys = list(rows_out[0].keys())
    header_row = 3
    for j, k in enumerate(keys, 1):
        cell = ws.cell(header_row, j, k)
        cell.font = Font(bold=True)
    for i, r in enumerate(rows_out, header_row + 1):
        for j, k in enumerate(keys, 1):
            ws.cell(i, j, r.get(k))


def _dict_rows_sheet(ws, rows: List[Dict[str, Any]], title: str) -> None:
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(bold=True)
    if not rows:
        ws.cell(2, 1, "(no rows)")
        return
    keys = list(rows[0].keys())
    hr = 3
    for j, k in enumerate(keys, 1):
        ws.cell(hr, j, k).font = Font(bold=True)
    for i, r in enumerate(rows, hr + 1):
        for j, k in enumerate(keys, 1):
            v = r.get(k)
            ws.cell(i, j, "" if v is None else v)


async def build_demo_seed_workbook_bytes() -> bytes:
    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    readme = wb.create_sheet("README", 0)
    readme.cell(1, 1, "Nigeria AML Compliance Platform — demo seed data export")
    readme.cell(1, 1).font = Font(bold=True)
    readme.cell(2, 1, f"Reference time anchor (UTC): {_EXPORT_REFERENCE_NOW.isoformat()}")
    lines = (
        "This workbook snapshots reference seed data used by demo API routes.",
        "Sheets: otc_branch_reference (10 STR/OTC intake rows), standard_aml_seed (POST /demo/seed),",
        "showcase_12_tracks (POST /demo/seed-showcase), ingest_flagship (POST /demo/ingest-flagship),",
        "temporal_profiles + temporal_scenarios (POST /demo/simulate-temporal / seed-complete-demo).",
        "Transaction ids are new UUIDs on each export; narrative/amounts/metadata match the code paths.",
        "Full 10-year temporal runs are procedural (generate_temporal_dataset); not exported row-by-row.",
        "AOP template seed metadata is environment-specific; see demo_aop_template_seed in backend.",
    )
    for i, line in enumerate(lines, 4):
        readme.cell(i, 1, line)

    otc_rows = export_reference_table_json_ready()
    otc_ws = wb.create_sheet("otc_branch_reference")
    _dict_rows_sheet(otc_ws, otc_rows, "Branch OTC / STR reference (10 rows)")

    ref_now = _EXPORT_REFERENCE_NOW
    standard: List[TransactionResponse] = []

    async def _collect_std(t: TransactionResponse) -> None:
        standard.append(t)

    await run_standard_demo_transaction_sequence(ref_now, _collect_std)

    std_ws = wb.create_sheet("standard_aml_seed")
    _transactions_to_sheet(std_ws, standard, "standard_aml_seed")

    showcase: List[TransactionResponse] = []

    async def _noop(_: TransactionResponse) -> None:
        return None

    await seed_high_risk_showcase(_noop, now=ref_now, capture=showcase)
    sh_ws = wb.create_sheet("showcase_12_tracks")
    _transactions_to_sheet(sh_ws, showcase, "showcase_12_tracks")

    flagship = _flagship_template(ref_now)
    fl_ws = wb.create_sheet("ingest_flagship")
    _transactions_to_sheet(fl_ws, [flagship], "ingest_flagship")

    prof_rows: List[Dict[str, Any]] = []
    for p in DEFAULT_PROFILES:
        prof_rows.append(
            {
                "customer_id": p.customer_id,
                "label": p.label,
                "salary_range": str(p.salary_range) if p.salary_range else "",
                "small_transfer_range": str(p.small_transfer_range),
                "max_normal_cash": p.max_normal_cash,
            }
        )
    tp_ws = wb.create_sheet("temporal_profiles")
    _dict_rows_sheet(tp_ws, prof_rows, "DEFAULT_PROFILES (temporal simulation)")

    scen_ws = wb.create_sheet("temporal_scenarios")
    scen_ws.cell(1, 1, "Scenario codes injected by generate_temporal_dataset()")
    scen_ws.cell(1, 1).font = Font(bold=True)
    for i, code in enumerate(_TEMPORAL_SCENARIO_CODES, 3):
        scen_ws.cell(i, 1, code)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
